import csv
import importlib
import io
import json
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db
from models.user import User
from models.transaction import Transaction
from services.auth_service import get_current_user
from services.ai_engine import detect_category
from services.gemini_service import generate_financial_summary

router = APIRouter(prefix="/api/data", tags=["Data Import/Export"])


@router.post("/import/csv")
async def import_csv(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Import transactions from a CSV file."""
    if not file.filename or not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail={"message": "Please upload a CSV file"})

    contents = await file.read()
    text = contents.decode("utf-8")
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = 0
    for row in reader:
        try:
            amount_str = row.get("amount", row.get("Amount", "0")).strip()
            amount = abs(float(amount_str.replace(",", "").replace("$", "").replace("₹", "").replace("£", "").replace("€", "")))
            if amount <= 0:
                skipped += 1
                continue
            description = row.get("description", row.get("Description", row.get("name", row.get("Name", "Imported")))).strip()
            date_str = row.get("date", row.get("Date", datetime.utcnow().strftime("%Y-%m-%d"))).strip()
            category = row.get("category", row.get("Category", "")).strip()
            txn_type = row.get("type", row.get("Type", "expense")).strip().lower()

            if not category:
                category = detect_category(description)

            if txn_type not in ("income", "expense"):
                txn_type = "expense"

            txn = Transaction(
                user_id=user.id,
                type=txn_type,
                amount=amount,
                category=category,
                description=description,
                date=date_str,
                notes="Imported from CSV",
            )
            db.add(txn)
            imported += 1
        except Exception:
            skipped += 1

    db.commit()
    return {"success": True, "imported": imported, "skipped": skipped, "total": imported + skipped}


@router.post("/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Import transactions from an Excel (.xlsx) file."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail={"message": "Please upload an Excel (.xlsx) file"})

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail={"message": "openpyxl not installed on server"})

    contents = await file.read()
    wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail={"message": "Excel file has no data rows"})

    # Build header map (case-insensitive)
    raw_headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    header_map = {h: i for i, h in enumerate(raw_headers)}

    imported = 0
    skipped = 0
    for row in rows[1:]:
        try:
            def cell(names, default=""):
                for n in names:
                    idx = header_map.get(n.lower())
                    if idx is not None and row[idx] is not None:
                        return str(row[idx]).strip()
                return default

            amount_str = cell(["amount", "total", "value"], "0")
            amount = abs(float(amount_str.replace(",", "").replace("$", "").replace("₹", "").replace("£", "").replace("€", "")))
            if amount <= 0:
                skipped += 1
                continue
            description = cell(["description", "name", "details", "merchant"], "Imported")
            date_str = cell(["date", "transaction date"], datetime.utcnow().strftime("%Y-%m-%d"))
            category = cell(["category", "type of expense"], "")
            txn_type = cell(["type", "transaction type"], "expense").lower()

            if not category:
                category = detect_category(description)
            if txn_type not in ("income", "expense"):
                txn_type = "expense"

            txn = Transaction(
                user_id=user.id,
                type=txn_type,
                amount=amount,
                category=category,
                description=description,
                date=date_str,
                notes="Imported from Excel",
            )
            db.add(txn)
            imported += 1
        except Exception:
            skipped += 1

    db.commit()
    wb.close()
    return {"success": True, "imported": imported, "skipped": skipped, "total": imported + skipped}


@router.post("/import/pdf")
async def import_pdf(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Import transactions from a PDF file by extracting text."""
    if not file.filename or not file.filename.endswith(".pdf"):
        raise HTTPException(status_code=400, detail={"message": "Please upload a PDF file"})

    import re

    contents = await file.read()
    text = ""

    # Try optional pdfplumber first (better text extraction in many cases), then pypdf.
    pdfplumber_module = None
    pypdf_reader = None

    try:
        pdfplumber_module = importlib.import_module("pdfplumber")
    except ImportError:
        pass

    try:
        pypdf_reader = importlib.import_module("pypdf").PdfReader
    except (ImportError, AttributeError):
        pass

    if pdfplumber_module is None and pypdf_reader is None:
        raise HTTPException(status_code=500, detail={"message": "PDF parsing library not installed (need pdfplumber or pypdf)"})

    if pdfplumber_module is not None:
        try:
            with pdfplumber_module.open(io.BytesIO(contents)) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except Exception:
            text = ""

    if not text and pypdf_reader is not None:
        try:
            reader = pypdf_reader(io.BytesIO(contents))
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        except Exception:
            raise HTTPException(status_code=400, detail={"message": "Could not extract text from PDF"})

    if not text.strip():
        raise HTTPException(status_code=400, detail={"message": "No readable text found in the PDF"})

    # Parse lines for transaction-like patterns:
    # Looking for lines with a date and an amount, e.g. "2026-01-15  Grocery Shopping  $45.00"
    date_pattern = re.compile(r"(\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{4})")
    amount_pattern = re.compile(r"[\$₹€£]?\s*(\d[\d,]*\.?\d*)")

    lines = text.split("\n")
    imported = 0
    skipped = 0

    for line in lines:
        line = line.strip()
        if not line or len(line) < 5:
            continue

        date_match = date_pattern.search(line)
        amounts = amount_pattern.findall(line)

        if date_match and amounts:
            try:
                date_str = date_match.group(1).replace("/", "-")
                # Use the last amount found (usually the total)
                amount = abs(float(amounts[-1].replace(",", "")))
                if amount == 0:
                    continue

                # Everything that's not date or amount is description
                desc = line
                desc = date_pattern.sub("", desc)
                desc = re.sub(r"[\$₹€£]?\s*\d[\d,]*\.?\d*", "", desc)
                desc = re.sub(r"\s+", " ", desc).strip(" -|,.")
                if not desc:
                    desc = "PDF Import"

                category = detect_category(desc)

                txn = Transaction(
                    user_id=user.id,
                    type="expense",
                    amount=amount,
                    category=category,
                    description=desc[:255],
                    date=date_str,
                    notes="Imported from PDF",
                )
                db.add(txn)
                imported += 1
            except Exception:
                skipped += 1

    db.commit()
    return {"success": True, "imported": imported, "skipped": skipped, "total": imported + skipped}


@router.get("/export/csv")
def export_csv(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export all transactions as CSV."""
    transactions = db.query(Transaction).filter(
        Transaction.user_id == user.id,
    ).order_by(Transaction.date.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Type", "Category", "Description", "Amount", "Notes"])
    for t in transactions:
        writer.writerow([t.date, t.type, t.category, t.description, t.amount, t.notes or ""])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=snapspend_export_{datetime.utcnow().strftime('%Y%m%d')}.csv"},
    )


@router.get("/export/json")
def export_json(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export all transactions as JSON."""
    transactions = db.query(Transaction).filter(
        Transaction.user_id == user.id,
    ).order_by(Transaction.date.desc()).all()

    data = {
        "exportDate": datetime.utcnow().isoformat(),
        "user": user.name,
        "transactionCount": len(transactions),
        "transactions": [t.to_dict() for t in transactions],
    }

    output = json.dumps(data, indent=2)
    return StreamingResponse(
        io.BytesIO(output.encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=snapspend_export_{datetime.utcnow().strftime('%Y%m%d')}.json"},
    )


@router.get("/export/excel")
def export_excel(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export all transactions as Excel spreadsheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail={"message": "openpyxl not installed"})

    transactions = db.query(Transaction).filter(
        Transaction.user_id == user.id,
    ).order_by(Transaction.date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    headers = ["Date", "Type", "Category", "Description", "Amount", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=t.date)
        ws.cell(row=row, column=2, value=t.type)
        ws.cell(row=row, column=3, value=t.category)
        ws.cell(row=row, column=4, value=t.description)
        ws.cell(row=row, column=5, value=t.amount)
        ws.cell(row=row, column=6, value=t.notes or "")

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=snapspend_export_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"},
    )


@router.get("/export/pdf")
def export_pdf(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export financial report as PDF."""
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail={"message": "fpdf2 not installed"})

    transactions = db.query(Transaction).filter(
        Transaction.user_id == user.id,
    ).order_by(Transaction.date.desc()).all()

    total_income = sum(t.amount for t in transactions if t.type == "income")
    total_expenses = sum(t.amount for t in transactions if t.type == "expense")
    savings = total_income - total_expenses

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Title
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 15, "SnapSpend AI - Financial Report", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, f"Generated: {datetime.utcnow().strftime('%B %d, %Y')}", ln=True, align="C")
    pdf.cell(0, 8, f"User: {user.name} ({user.email})", ln=True, align="C")
    pdf.ln(10)

    # Summary
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Financial Summary", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Total Income: ${total_income:,.2f}", ln=True)
    pdf.cell(0, 8, f"Total Expenses: ${total_expenses:,.2f}", ln=True)
    pdf.cell(0, 8, f"Net Savings: ${savings:,.2f}", ln=True)
    pdf.cell(0, 8, f"Savings Rate: {round(savings / max(total_income, 1) * 100, 1)}%", ln=True)
    pdf.cell(0, 8, f"Transactions: {len(transactions)}", ln=True)
    pdf.ln(8)

    # Category breakdown
    categories = {}
    for t in transactions:
        if t.type == "expense":
            categories[t.category] = categories.get(t.category, 0) + t.amount

    if categories:
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Category Breakdown", ln=True)
        pdf.set_font("Helvetica", "", 10)
        for cat, amount in sorted(categories.items(), key=lambda x: x[1], reverse=True):
            pct = round(amount / max(total_expenses, 1) * 100, 1)
            pdf.cell(0, 7, f"  {cat.title()}: ${amount:,.2f} ({pct}%)", ln=True)
        pdf.ln(5)

    # Transactions table
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Transaction Details", ln=True)

    # Table headers
    pdf.set_font("Helvetica", "B", 9)
    col_widths = [25, 20, 25, 70, 25, 25]
    headers = ["Date", "Type", "Category", "Description", "Amount", "Notes"]
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 8, header, border=1)
    pdf.ln()

    # Table rows (limited to 50 recent)
    pdf.set_font("Helvetica", "", 8)
    for t in transactions[:50]:
        pdf.cell(col_widths[0], 7, t.date[:10], border=1)
        pdf.cell(col_widths[1], 7, t.type, border=1)
        pdf.cell(col_widths[2], 7, t.category[:10], border=1)
        pdf.cell(col_widths[3], 7, (t.description or "")[:35], border=1)
        pdf.cell(col_widths[4], 7, f"${t.amount:,.2f}", border=1)
        pdf.cell(col_widths[5], 7, (t.notes or "")[:12], border=1)
        pdf.ln()

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=snapspend_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf"},
    )


@router.get("/export/ai-summary")
def export_ai_summary(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get an AI-generated financial summary text."""
    transactions = db.query(Transaction).filter(
        Transaction.user_id == user.id,
    ).all()

    total_income = sum(t.amount for t in transactions if t.type == "income")
    total_expenses = sum(t.amount for t in transactions if t.type == "expense")

    categories = {}
    for t in transactions:
        if t.type == "expense":
            categories[t.category] = categories.get(t.category, 0) + t.amount

    transactions_text = f"""
Total Income: ${total_income:,.2f}
Total Expenses: ${total_expenses:,.2f}
Net Savings: ${total_income - total_expenses:,.2f}
Transaction Count: {len(transactions)}
Top Categories: {', '.join(f'{k}: ${v:,.2f}' for k, v in sorted(categories.items(), key=lambda x: x[1], reverse=True)[:5])}
"""

    ai_summary = generate_financial_summary(transactions_text)
    if not ai_summary:
        # Fallback
        savings_rate = round((total_income - total_expenses) / max(total_income, 1) * 100, 1)
        top_cats = sorted(categories.items(), key=lambda x: x[1], reverse=True)[:3]
        ai_summary = f"""Financial Summary for {user.name}

Total Income: ${total_income:,.2f} | Total Expenses: ${total_expenses:,.2f}
Net Savings: ${total_income - total_expenses:,.2f} ({savings_rate}% savings rate)

Top spending categories: {', '.join(f'{cat.title()} (${amt:,.2f})' for cat, amt in top_cats)}.

{'Great job maintaining a healthy savings rate!' if savings_rate > 20 else 'Consider reducing discretionary spending to improve your savings rate.'}"""

    return {"summary": ai_summary}
